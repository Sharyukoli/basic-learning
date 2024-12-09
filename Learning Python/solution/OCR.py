import os
import time
from openpyxl import load_workbook 
import slate3k as slate
import tkinter as tk
from tkinter import filedialog
import datetime
from datetime import datetime
from datetime import date
import tkinter.font as tkFont
import tkcalendar as tkcall
from tkcalendar import Calendar as tkCalendar
import calendar as call
import sys
import pandas as pd
from tkinter import messagebox
global folder_path
global excel_file
global selected_date
global pdf_count
global po_num
global request_date
global pur_order_date
global item_no
global ship_to
global ncm
global material_desc
global order_qty
global unit_price

root = tk.Tk()
#setting title
root.title("Get Scanned P.Os to Excel")
#setting window size
width=561
height=433
screenwidth = root.winfo_screenwidth()
screenheight = root.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
root.geometry(alignstr)
root.resizable(width=False, height=False)

def Open_Folder_Button_command():
    print('inside command')
    global folder_path 
    folder_path = filedialog.askdirectory()
    print(folder_path)
    if folder_path:
        print(folder_path)
        
    print(" outside command")

def close_window():
    root.destroy()


def cancel_process():
    result = tk.messagebox.askyesno("Cancel Process", "Are you sure you want to cancel the process?")
    if result:
        sys.exit()

def open_folder():
    global folder_path
    folder_path = filedialog.askdirectory()
    if folder_path:
        Open_Folder_Label.delete(0, tk.END)
        Open_Folder_Label.insert(0,folder_path)
    print('FOlder selected is= ',folder_path)
    
def on_process(file_path):    
    global pdf_count
    pdf_count=0
    file_count = 0
    pdf_files = []
    for file in os.listdir(folder_path):
        if file.endswith(".pdf"):
            pdf_file_path = os.path.join(folder_path, file)
            downloaded_time = datetime.fromtimestamp(os.path.getmtime(pdf_file_path))
            global selected_date
            selected_date = cal.get_date()
            selected_date = selected_date.strftime("%d/%m/%Y")
            new_selected_date = datetime.strptime(selected_date, "%d/%m/%Y")  ## Convert calendar date string to date
            new_selected_date = new_selected_date.strftime("%d/%m/%Y")
            new_downloaded_time = downloaded_time.strftime("%d/%m/%Y")
            
            if new_downloaded_time == new_selected_date:
                print('Hiiiiiiiiiiiiiiiiiiii')
                pdf_file = open(os.path.join(folder_path, file), 'rb')
                pdf_count += 1
                pdf_files.append(file)
                print(file)
                extracted_text=slate.PDF(pdf_file)
                new_text = extracted_text[0]
                text = " ".join(extracted_text[0].split())
                col_dict = {
                "D": "Purchase Order P.O. Number: ",
                "E": "Date: ",
                "Y": "Contact Person:",
                "G": "Bill To ",
                "Z" : "Terms of Payment:",
                "H": "Item Material N.C.M. Description Material Use *Vend # / ** Mf #",
                "K": "Order Quantity Units ",
                "Q": "*Vend # / ** Mf # ",
                "R": "Item Value Tax (%) ",
                "N":"Page Number:"
                }
                pos = []
                for values in col_dict.values():
                    find_pos = text.find(values)
                    pos.append(find_pos)
                global po_num
                po_num = text[pos[0]+len(col_dict["D"]):pos[1]]
                print('Po num= ', po_num)
                ####
                global pur_order_date
                pur_order_date = text[pos[1]+len(col_dict["E"]):pos[2]]
                print("pur date= ", pur_order_date)
                ####
                global ship_to
                ship_to = text[pos[3]+len(col_dict["G"]):pos[4]]
                print("Ship to= ", ship_to)
                ####

                abc = text[pos[5]+len(col_dict["H"]):pos[6]]
                abc = abc.split(" ")

                global item_no
                item_no = abc[1]
                print("Item_no=", item_no)
                #####
                
                global material_desc
                material_desc = ""
                for words in abc:
                    if words.isupper() == True:
                        #print(words)
                        material_desc = material_desc + " " + words
                print("MATERIAL DESC===== ", material_desc)
                #####

                global ncm
                ncm = abc[2]
                print('ncm = ', ncm)
                #####

                global order_qty
                xyz = text[pos[6]+len(col_dict["K"]):pos[8]].split(" ")
                order_qty=xyz[0]
                #####
                print('qty= ',xyz[0])

                global unit_price
                efg = text[pos[8]:pos[9]].split(" ")
                #print('Efg =',efg)
                unit_price = efg[-4]

                global request_date
                request_date = efg[8]
                print('unit_price =', unit_price)
                #####

                total_price = (efg[-2])
                print('total_price=', total_price)
                print("REMOVEEEEEEEEEEEEEEEEEEEE")
                #print(text)

                wb = load_workbook(file_path)
                sheet = wb["Sheet1"]
                row_count = sheet.max_row
    
                
                # GLabel_941.delete(0,tk.END)              
                # GLabel_941.insert(0,row_count)
                #sheet = workbook.active
                sheet.insert_rows(row_count+1)
                rows_updae = 0
                for col in range(1, sheet.max_column + 1):
                    #sheet.column_dimensions["AA"].hidden = False
                    #request_date
                    if col == 3:
                        print('inside column 4')
                        sheet.cell(row=row_count+1, column=col, value=request_date)
                    elif col == 4:
                        print('inside column 4')
                        sheet.cell(row=row_count+1, column=col, value=po_num)
                    elif col== 5:
                        sheet.cell(row=row_count+1, column=col, value=pur_order_date)
                    elif col== 17:
                        sheet.cell(row=row_count+1, column=col, value=material_desc)
                    elif col== 9:
                        sheet.cell(row=row_count+1, column=col, value=item_no)
                    
                    elif col== 11:
                        sheet.cell(row=row_count+1, column=col, value=order_qty)
                    elif col== 18:
                        sheet.cell(row=row_count+1, column=col, value=unit_price)
                    elif col== 28:
                        print('inside 288888888888888888888888888888888')
                        print(file)
                        sheet.cell(row=row_count+1, column=col, value=file)

                    else:
                        pass
                    #sheet.column_dimensions["AA"].hidden = True
                    
                    wb.save(file_path)
                    #rows_updae +=1
                    #GLabel_262.delete(0,tk.END)
                    #GLabel_262.insert(0,rows_updae)
            file_count += 1
            GLabel_17.delete(0,tk.END)              
            GLabel_17.insert(0,file_count)
            
                    
           

            
    GLabel_590.delete(0,tk.END)              
    GLabel_590.insert(0,pdf_count)                   # Number of PDF files
    GLabel_262.delete(0,tk.END)
    GLabel_262.insert(0,pdf_count)
    messagebox.showinfo("Success", "File edited successfully")


    print(pdf_files)


def select_pdf():
    for file in os.listdir(folder_path):
        if file.endswith(".pdf"):
            print(file)
    print("command")


def select_excel():
    global excel_file
    #excel_file = filedialog.askopenfilename(filetypes=[("Excel file",'.xlsx')])
    file_path = filedialog.askopenfilename(initialdir = "/", title = "Select file", filetypes = (("Excel files", "*.xlsx"), ("all files", "*")))
    
    
    if file_path:
        wb = load_workbook(file_path)
        sheet = wb["Sheet1"]
        row_count = sheet.max_row
        GLabel_941.delete(0,tk.END)              
        GLabel_941.insert(0,row_count)
        excel_file = file_path
        excel_path = str(excel_file)
        # print('EXCELLLLLLLLL PATH',excel_path)
        # print('EXCELLLLLLLLLL FILEEEEEEE: ', excel_file)
        index_i = excel_file.rfind('/')
        excel_file_name = excel_file[index_i+1:]
        print(excel_file_name)
        Selected_Excel.delete(0, tk.END)
        Selected_Excel.insert(0,excel_file_name)
        on_process(file_path)
    else:
        messagebox.showerror("Error", "No file selected")
    
    
    

def open_excel_file():
    os.startfile(excel_file)
    print("excel opened")

############### Window ICON #####################  ↓

p1 = tk.PhotoImage(file=r"D:\ETSS_Project\sfinal.png")
root.iconphoto(False, p1)

################# LOGO  #########################  ↓

f_img = tk.PhotoImage(file=r"D:\ETSS_Project\sfinal.png")
panel1 = tk.Label(image=f_img)
panel1.image = f_img
panel1.place(x=7, y=250, width=160, height=90)

############## CALENDAR DROPDOWN #################  ↓

todays_date = date.today()
cal = tkcall.DateEntry(root,x=30,y=100, width=12, selectmode='day',date_pattern='dd/MM/yyyy', year=todays_date.year, month=todays_date.month, day=todays_date.day, background='darkblue', foreground='white',borderwidth=1)
cal.pack(padx=230,pady=30,anchor=tk.W)
selected_date = cal.get_date()
selected_date = selected_date.strftime("%d/%m/%Y")
cal.delete(0, tk.END)
cal.insert(0,selected_date)
print(selected_date)

###################### GUI ######################

Date_Label=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
Date_Label["font"] = ft
Date_Label["fg"] = "#333333"
Date_Label["justify"] = "center"
Date_Label["text"] = "Enter Download Date of PDF file"
Date_Label.place(x=20,y=30,width=180,height=30)                     ###  DATE / CALENDAR LABEL

##
Calendar_Label=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
Calendar_Label["font"] = ft
Calendar_Label["fg"] = "#333333"
Calendar_Label["justify"] = "center"
Calendar_Label["text"] = "Choose the Folder"
Calendar_Label.place(x=20,y=65,width=102,height=30)                 ### FOLDER LABEL
##
GLabel_974=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_974["font"] = ft
GLabel_974["fg"] = "#333333"
GLabel_974["justify"] = "center"
GLabel_974["text"] = "Choose the Excel Sheet"
GLabel_974.place(x=20,y=100,width=132,height=30)                   ## EXCEL SHEET LABEL


Open_Folder_Label=tk.Entry(root)
Open_Folder_Label["borderwidth"] = "1px"
ft = tkFont.Font(family='Times',size=10)
Open_Folder_Label["font"] = ft
Open_Folder_Label["fg"] = "#333333"
Open_Folder_Label["justify"] = "center"
Open_Folder_Label["text"] = ""
Open_Folder_Label.place(x=230,y=65,width=150,height=23)             ### DISPLAY Folder 


Open_Folder_Button=tk.Button(root)
Open_Folder_Button["bg"] = "#f0f0f0"
ft = tkFont.Font(family='Times',size=10)
Open_Folder_Button["font"] = ft
Open_Folder_Button["fg"] = "#000000"
Open_Folder_Button["justify"] = "center"
Open_Folder_Button["text"] = "Browse"
Open_Folder_Button.place(x=410,y=60,width=70,height=25)
Open_Folder_Button["command"] = open_folder                  ### Choose Folder BUTTON


Selected_Excel=tk.Entry(root)
Selected_Excel["borderwidth"] = "1px"
ft = tkFont.Font(family='Times',size=10)
Selected_Excel["font"] = ft
Selected_Excel["fg"] = "#333333"
Selected_Excel["justify"] = "left"
Selected_Excel["text"] = ""
Selected_Excel.place(x=230,y=100,width=150,height=23)              ##Excel sheet

Open_Excel_Button=tk.Button(root)
Open_Excel_Button["bg"] = "#f0f0f0"
ft = tkFont.Font(family='Times',size=10)
Open_Excel_Button["font"] = ft
Open_Excel_Button["fg"] = "#000000"
Open_Excel_Button["justify"] = "center"
Open_Excel_Button["text"] = "Browse"
Open_Excel_Button.place(x=410,y=100,width=70,height=25)
Open_Excel_Button["command"] = select_excel                   ##Excel sheet BUTTON

GListBox_94=tk.Listbox(root)
GListBox_94["borderwidth"] = "1px"
ft = tkFont.Font(family='Times',size=10)
GListBox_94["font"] = ft
GListBox_94["bg"] = "#F0F0F0"
GListBox_94["justify"] = "center"
GListBox_94.place(x=170,y=170,width=308,height=190)

GLabel_142=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_142["font"] = ft
GLabel_142["fg"] = "#333333"
GLabel_142["justify"] = "center"
GLabel_142["text"] = "Number of PDF Files : "
GLabel_142.place(x=190,y=190,width=123,height=30)

GLabel_36=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_36["font"] = ft
GLabel_36["fg"] = "#333333"
GLabel_36["justify"] = "center"
GLabel_36["text"] = "Number of Files Processed :"
GLabel_36.place(x=190,y=230,width=157,height=30)

GLabel_804=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_804["font"] = ft
GLabel_804["fg"] = "#333333"
GLabel_804["justify"] = "center"
GLabel_804["text"] = "Number of Existing rows in Excel : "
GLabel_804.place(x=190,y=270,width=190,height=30)

GLabel_515=tk.Label(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_515["font"] = ft
GLabel_515["fg"] = "#333333"
GLabel_515["justify"] = "center"
GLabel_515["text"] = "Number of  new rows created :"
GLabel_515.place(x=190,y=310,width=170,height=30)

GLabel_17=tk.Entry(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_17["font"] = ft
GLabel_17["bg"] = "#FFFFFF"
GLabel_17["justify"] = "center"
GLabel_17["text"] = ""
GLabel_17.place(x=390,y=190,width=70,height=25)

GLabel_590=tk.Entry(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_590["font"] = ft
GLabel_590["bg"] = "#FFFFFF"
GLabel_590["justify"] = "center"
GLabel_590["text"] = "2 file"
GLabel_590.place(x=390,y=230,width=70,height=25)

GLabel_941=tk.Entry(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_941["font"] = ft
GLabel_941["bg"] = "#FFFFFF"
GLabel_941["justify"] = "center"
GLabel_941["text"] = "3 ex ro"
GLabel_941.place(x=390,y=270,width=70,height=25)

GLabel_262=tk.Entry(root)
ft = tkFont.Font(family='Times',size=10)
GLabel_262["font"] = ft
GLabel_262["bg"] = "#FFFFFF"
GLabel_262["justify"] = "center"
GLabel_262["text"] = "4 new"
GLabel_262.place(x=390,y=310,width=70,height=25)

GButton_203=tk.Button(root)
GButton_203["bg"] = "#f0f0f0"
ft = tkFont.Font(family='Times',weight='bold',size=10)
GButton_203["font"] = ft
GButton_203["fg"] = "#000000"
GButton_203["justify"] = "center"
GButton_203["text"] = "Open Excel"
GButton_203.place(x=190,y=380,width=70,height=25)
GButton_203["command"] = open_excel_file

GButton_492=tk.Button(root)
GButton_492["bg"] = "#f0f0f0"
ft = tkFont.Font(family='Times',weight='bold',size=10)
GButton_492["font"] = ft
GButton_492["fg"] = "#000000"
GButton_492["justify"] = "center"
GButton_492["text"] = "Process"
GButton_492.place(x=300,y=380,width=70,height=25)
GButton_492["command"] = on_process

GButton_395=tk.Button(root)
GButton_395["bg"] = "#f0f0f0"
ft = tkFont.Font(family='Times',weight='bold',size=10)
GButton_395["font"] = ft
GButton_395["fg"] = "#000000"
GButton_395["justify"] = "center"
GButton_395["text"] = "Cancel"
GButton_395.place(x=410,y=380,width=70,height=25)
GButton_395["command"] = cancel_process

GButton_331=tk.Button(root)
GButton_331["bg"] = "#f0f0f0"
ft = tkFont.Font(family='Times',weight='bold',size=10)
GButton_331["font"] = ft
GButton_331["fg"] = "#FF0000"
GButton_331["justify"] = "center"
GButton_331["text"] = "Close"
GButton_331.place(x=35,y=380,width=70,height=25)
GButton_331["command"] = close_window

root.mainloop()
